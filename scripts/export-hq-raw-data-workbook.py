#!/usr/bin/env python3
"""Export all HQ raw data lake pulls into one Excel workbook.

Usage:
  python3 scripts/export-hq-raw-data-workbook.py

Inputs:
  docs/data/impact_coverage_all_reps.json  — sql/18
  docs/data/rep_jv_all_reps.json           — sql/19
  docs/data/pcid_market_attributes.json    — sql/20

Output:
  docs/data/hq_raw_data.xlsx
    - Impact coverage
    - JV / JAM
    - PCID Market
    - About
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
IC_PATH = ROOT / "docs" / "data" / "impact_coverage_all_reps.json"
JV_PATH = ROOT / "docs" / "data" / "rep_jv_all_reps.json"
REP_BOOK_PATH = ROOT / "docs" / "data" / "rep_book.json"
PCID_JSON_PATH = ROOT / "docs" / "data" / "pcid_market_attributes.json"
PCID_CSV_PATH = ROOT / "docs" / "data" / "pcid_market_attributes.csv"
OUT_PATH = ROOT / "docs" / "data" / "hq_raw_data.xlsx"


def load_json(path: Path, required: bool = True) -> dict:
    if not path.is_file():
        if required:
            print(f"Missing input: {path}")
            sys.exit(1)
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def load_csv_rows(path: Path) -> list[dict]:
    import csv

    if not path.is_file():
        return []
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def rows_from_payload(payload: dict, key: str = "reps") -> list[dict]:
    if key in payload:
        return payload[key]
    if "rows" in payload:
        return payload["rows"]
    if "data" in payload:
        return payload["data"]
    return []


def load_pcid(pcid_json: dict) -> tuple[list[dict], dict]:
    rows = rows_from_payload(pcid_json, "rows")
    if rows:
        return rows, pcid_json
    rows = load_csv_rows(PCID_CSV_PATH)
    if rows:
        meta = dict(pcid_json) if pcid_json else {}
        meta.setdefault("query", "sql/20_pcid_market_attributes.sql")
        meta.setdefault("source_table", "datalake.scss.client_attributes_dim_parent_attributes_current")
        meta["row_count"] = len(rows)
        return rows, meta
    print(f"Missing PCID data: {PCID_JSON_PATH} or {PCID_CSV_PATH}")
    sys.exit(1)


def column_order(rows: list[dict]) -> list[str]:
    if not rows:
        return []
    keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for k in row:
            if k is None or k in seen:
                continue
            seen.add(k)
            keys.append(k)
    return keys


def write_sheet(ws, rows: list[dict], keys: list[str] | None = None) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    keys = keys or column_order(rows)
    labels = [(k or "column").replace("_", " ").title() for k in keys]
    ws.append(labels)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(k) for k in keys])
    for col_idx, label in enumerate(labels, start=1):
        width = min(max(len(str(label)) + 2, 12), 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def about_rows(
    ic: dict,
    jv: dict,
    pcid: dict,
) -> list[tuple[str, str]]:
    return [
        ("Workbook", "HQ raw data lake export"),
        ("Generated", date.today().isoformat()),
        ("", ""),
        ("Sheet: Impact coverage", ""),
        ("Query", ic.get("query", "sql/18_impact_coverage_all_reps.sql")),
        ("Execution ID", ic.get("execution_id", "")),
        ("Updated at", ic.get("updated_at", "")),
        ("Row count", str(ic.get("row_count", len(rows_from_payload(ic))))),
        ("Source tables", ", ".join(ic.get("source_tables", []))),
        ("Window", ic.get("window", "")),
        ("Note", ic.get("note", "")),
        ("", ""),
        ("Sheet: JV / JAM", ""),
        ("Query", jv.get("query", "sql/19_rep_jv_all_reps.sql")),
        ("Execution ID", jv.get("execution_id", "")),
        ("Updated at", jv.get("updated_at", "")),
        ("Row count", str(jv.get("row_count", len(rows_from_payload(jv))))),
        ("Source tables", ", ".join(jv.get("source_tables", [
            "datalake.imhotep_iceberg.jobactivitymetrics",
            "datalake.sales_data_strategy_dsa.current_parent_rep_assignment",
        ]))),
        ("Note", jv.get("note", "jobs_90d, rev_per_job, revenue_90d, pqr_90d by rep")),
        ("", ""),
        ("Sheet: PCID Market", ""),
        ("Query", pcid.get("query", "sql/20_pcid_market_attributes.sql")),
        ("Execution ID", ", ".join(pcid.get("execution_ids", [pcid.get("execution_id", "")]))),
        ("Updated at", pcid.get("updated_at", "")),
        ("Row count", str(pcid.get("row_count", len(rows_from_payload(pcid, "rows"))))),
        ("Source table", pcid.get("source_table", "datalake.scss.client_attributes_dim_parent_attributes_current")),
        ("Scope note", pcid.get("note", "")),
    ]


def jv_by_bucket_rows() -> list[dict]:
    """Market × PCID bucket median $/job from rep_book + rep_jv (sql/21 logic)."""
    sys.path.insert(0, str(ROOT / "scripts"))
    from build_market_summary import (
        compute_jv_by_bucket,
        load_rep_book_by_market,
        load_rep_jv_by_id,
        merge_reps_with_jv,
    )

    rep_book_by_market = load_rep_book_by_market(REP_BOOK_PATH)
    jv_by_id = load_rep_jv_by_id(JV_PATH)
    rows: list[dict] = []
    for market_key, reps in sorted(rep_book_by_market.items()):
        merged = merge_reps_with_jv(reps, jv_by_id)
        if not merged:
            continue
        country, segment = market_key.split("-", 1) if "-" in market_key else ("", market_key)
        for bucket in compute_jv_by_bucket(merged):
            rows.append(
                {
                    "market": market_key,
                    "country": country,
                    "segment": segment,
                    **bucket,
                }
            )
    return rows


def write_xlsx(ic: dict, jv: dict, pcid: dict, path: Path) -> dict[str, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("openpyxl not installed — pip install openpyxl")
        sys.exit(1)

    ic_rows = rows_from_payload(ic)
    jv_rows = rows_from_payload(jv)
    pcid_rows = rows_from_payload(pcid, "pcids") or rows_from_payload(pcid, "rows")

    wb = Workbook()
    ws_ic = wb.active
    ws_ic.title = "Impact coverage"
    write_sheet(ws_ic, ic_rows)

    ws_jv = wb.create_sheet("JV JAM")
    write_sheet(ws_jv, jv_rows)

    jv_bucket_rows = jv_by_bucket_rows()
    ws_jv_bucket = wb.create_sheet("JV by bucket")
    write_sheet(ws_jv_bucket, jv_bucket_rows)

    ws_pcid = wb.create_sheet("PCID Market")
    write_sheet(ws_pcid, pcid_rows)

    ws_about = wb.create_sheet("About")
    bold = Font(bold=True)
    ws_about.append(["Field", "Value"])
    for cell in ws_about[1]:
        cell.font = bold
    for field, value in about_rows(ic, jv, pcid):
        ws_about.append([field, value])
    ws_about.column_dimensions["A"].width = 28
    ws_about.column_dimensions["B"].width = 80

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    counts = {
        "Impact coverage": len(ic_rows),
        "JV JAM": len(jv_rows),
        "JV by bucket": len(jv_bucket_rows),
        "PCID Market": len(pcid_rows),
    }
    print(f"Wrote {path}")
    for sheet, count in counts.items():
        print(f"  {sheet}: {count:,} rows")
    return counts


def main() -> None:
    ic = load_json(IC_PATH)
    jv = load_json(JV_PATH)
    pcid_json = load_json(PCID_JSON_PATH, required=False)
    pcid_rows, pcid_meta = load_pcid(pcid_json)
    write_xlsx(ic, jv, {**pcid_meta, "rows": pcid_rows}, OUT_PATH)


if __name__ == "__main__":
    main()
