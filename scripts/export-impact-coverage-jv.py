#!/usr/bin/env python3
"""Export impact coverage + JV (job value) for all reps to a single Excel sheet.

Usage:
  python3 scripts/export-impact-coverage-jv.py

Inputs:
  docs/data/impact_coverage_all_reps.json  — impact coverage (sql/18)
  docs/data/rep_jv_all_reps.json           — jobs_90d, rev_per_job (sql/19, optional)
  docs/data/headcount.json                 — segment opp_plateau_rev_per_job (sql/16)

Output:
  docs/data/impact_coverage_jv.xlsx
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
IC_PATH = ROOT / "docs" / "data" / "impact_coverage_all_reps.json"
JV_PATH = ROOT / "docs" / "data" / "rep_jv_all_reps.json"
HC_PATH = ROOT / "docs" / "data" / "headcount.json"
OUT_PATH = ROOT / "docs" / "data" / "impact_coverage_jv.xlsx"

FIELD_LABELS: dict[str, str] = {
    "sales_rep_id": "Sales rep ID",
    "country": "Country",
    "segment": "Segment",
    "sales_team_name": "Sales team",
    "pcid_count": "PCID count",
    "impact_calls_90d": "Impact calls 90d",
    "impact_calls_per_account": "Impact calls / account",
    "segment_avg_coverage": "Segment avg coverage",
    "jobs_90d": "Jobs 90d",
    "rev_per_job": "Rev / job ($)",
    "opp_plateau_rev_per_job": "Opp plateau $/job (segment)",
    "pqr_90d": "PQR 90d ($)",
    "revenue_90d": "Revenue 90d ($)",
    "segment_avg_pcid": "Segment avg PCID",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "too_big_coverage_signal": "Too big (coverage signal)",
}

KEY_ORDER = list(FIELD_LABELS.keys())


def load_json(path: Path) -> dict:
    if not path.is_file():
        print(f"Missing input: {path}")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def load_json_optional(path: Path) -> dict:
    if path.is_file():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def segment_jv_lookup(headcount: dict) -> dict[tuple[str, str], float | None]:
    lookup: dict[tuple[str, str], float | None] = {}
    for m in headcount.get("markets", []):
        key = (m.get("country", ""), m.get("segment", ""))
        lookup[key] = m.get("opp_plateau_rev_per_job")
    return lookup


def jv_lookup(jv_payload: dict) -> dict[tuple[str, str, int], dict]:
    lookup: dict[tuple[str, str, int], dict] = {}
    for rep in jv_payload.get("reps", []):
        key = (rep["country"], rep["segment"], int(rep["sales_rep_id"]))
        lookup[key] = rep
    return lookup


def merge_rows(ic_payload: dict, jv_payload: dict, seg_jv: dict[tuple[str, str], float | None]) -> list[dict]:
    jv_by_rep = jv_lookup(jv_payload) if jv_payload else {}
    rows: list[dict] = []
    for rep in ic_payload.get("reps", []):
        row = dict(rep)
        key = (row["country"], row["segment"], int(row["sales_rep_id"]))
        jv = jv_by_rep.get(key, {})
        row["jobs_90d"] = jv.get("jobs_90d")
        row["rev_per_job"] = jv.get("rev_per_job")
        if jv.get("rev_per_job") is None and row.get("jobs_90d") and row.get("revenue_90d"):
            jobs = row["jobs_90d"]
            if jobs:
                row["rev_per_job"] = round(float(row["revenue_90d"]) / float(jobs), 2)
        row["opp_plateau_rev_per_job"] = seg_jv.get((row["country"], row["segment"]))
        rows.append(row)
    return rows


def label_for(key: str) -> str:
    return FIELD_LABELS.get(key, key.replace("_", " ").title())


def write_xlsx(rows: list[dict], ic_payload: dict, jv_payload: dict, path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    bold = Font(bold=True)
    ws = wb.active
    ws.title = "Impact coverage + JV"

    keys = KEY_ORDER
    ws.append([label_for(k) for k in keys])
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(k) for k in keys])

    for col_idx, key in enumerate(keys, start=1):
        label = label_for(key)
        width = min(max(len(str(label)) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions

    ws_meta = wb.create_sheet("About")
    ws_meta.append(["Field", "Value"])
    for cell in ws_meta[1]:
        cell.font = bold
    meta = [
        ("Impact coverage as of", ic_payload.get("updated_at", "")),
        ("Impact coverage query", ic_payload.get("query", "")),
        ("JV data as of", jv_payload.get("updated_at", "")),
        ("JV query", jv_payload.get("query", "")),
        ("Revenue window", ic_payload.get("window", "")),
        ("Note", "rev_per_job = revenue_90d / jobs_90d. opp_plateau_rev_per_job is segment-level from sql/16."),
        ("Export generated", date.today().isoformat()),
        ("Row count", len(rows)),
    ]
    for label, value in meta:
        ws_meta.append([label, value])
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 72

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {len(rows)} rows to {path}")


def main() -> None:
    ic_payload = load_json(IC_PATH)
    jv_payload = load_json_optional(JV_PATH)
    headcount = load_json_optional(HC_PATH)
    seg_jv = segment_jv_lookup(headcount) if headcount else {}
    rows = merge_rows(ic_payload, jv_payload, seg_jv)
    write_xlsx(rows, ic_payload, jv_payload, OUT_PATH)


if __name__ == "__main__":
    main()
