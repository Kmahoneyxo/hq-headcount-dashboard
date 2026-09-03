#!/usr/bin/env python3
"""Simulate ReferenceCheck.gs warehouse key matching + merge for Rep_Level rollups."""

import json
import re
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parents[1]
HC = ROOT / "docs" / "data" / "headcount.json"
GS = ROOT / "docs" / "google-apps-script" / "ReferenceCheck.gs"

SEGMENT_ALIASES = {
    "MEDIUM": "M", "MID": "M", "MED": "M", "MICRO": "M",
    "LARGE": "L", "SMALL": "NAM", "ENTERPRISE": "ACC",
    "ACCOUNT": "ACC", "ACCOUNTS": "ACC",
}
KNOWN = ["M", "UMM", "ACC", "L", "NAM", "DCA", "ISDCA", "NAMDCA"]


def normalize_segment(seg: str) -> str:
    s = str(seg or "").strip().upper()
    return SEGMENT_ALIASES.get(s, s)


def rollup_country(c: str) -> str:
    c = str(c or "").strip().upper()
    if c == "GB":
        return "UK"
    if c in ("IE", "IRELAND"):
        return "IRELAND"
    if c in ("DE", "AT", "CH"):
        return "DACH"
    if c in ("BE", "NL", "LU"):
        return "BNL"
    if c in ("ES", "PT"):
        return "IBE"
    return c


def segment_market_key(country: str, segment: str) -> Optional[str]:
    c = rollup_country(country)
    seg = normalize_segment(segment)
    if not c or not seg:
        return None
    return f"{c}-{seg}"


def parse_market_key(key: str) -> Optional[dict]:
    m = re.match(r"^([A-Z]{2,8})-([A-Z]+)$", str(key or "").strip().upper())
    if not m:
        return None
    return {"country": rollup_country(m.group(1)), "segment": m.group(2)}


def warehouse_key_aliases(country: str, segment: str, market: Optional[str] = None) -> list:
    keys = []
    seg = normalize_segment(segment)
    primary = segment_market_key(country, seg)
    if primary:
        keys.append(primary)
    c = rollup_country(country)
    if c == "IRELAND":
        keys.append(f"IE-{seg}")
    if c == "IE":
        keys.append(f"IRELAND-{seg}")
    if market:
        mk = str(market).strip().upper().replace(" ", "")
        if mk and mk not in keys:
            keys.append(mk)
        parsed = parse_market_key(mk)
        if parsed:
            alt = segment_market_key(parsed["country"], parsed["segment"])
            if alt and alt not in keys:
                keys.append(alt)
    return keys


def build_by_key(markets: list[dict]) -> dict[str, dict]:
    by_key: dict[str, dict] = {}
    for m in markets:
        for k in warehouse_key_aliases(m["country"], m["segment"], m.get("market")):
            if k and k not in by_key:
                by_key[k] = m
    return by_key


def lookup(by_key: dict, country: str, segment: str, market: str) -> Optional[dict]:
    for k in warehouse_key_aliases(country, segment, market):
        if k in by_key:
            return by_key[k]
    return None


def load_embedded() -> dict[str, list]:
    text = GS.read_text()
    m = re.search(r"var EMBEDDED_WAREHOUSE_METRICS = (\{.*?\});", text, re.DOTALL)
    if not m:
        raise SystemExit("EMBEDDED_WAREHOUSE_METRICS not found")
    return json.loads(m.group(1))


def expand_embedded(compact: dict) -> list[dict]:
    fields = [
        "revenue_90d", "avg_pqr_per_rep", "segment_avg_pqr", "segment_avg_pcid",
        "coverage_peak_accounts", "median_impact_calls_per_account", "coverage_at_inflection",
    ]
    markets = []
    for key, arr in compact.items():
        parts = parse_market_key(key)
        if not parts:
            continue
        row = {"country": parts["country"], "segment": parts["segment"], "market": key}
        for i, f in enumerate(fields):
            if arr[i] is not None:
                row[f] = arr[i]
        markets.append(row)
    return markets


def rep_rollups_from_xlsx() -> list[dict]:
    xlsx = Path("/Users/kmahoney/Downloads/Global Sales Rep Headcount (1).xlsx")
    if not xlsx.exists():
        return []
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Rep_Level"]
    header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    hmap = {str(h).strip().lower(): i for i, h in enumerate(header) if h}
    mi = hmap.get("market")
    ti = hmap.get("team_name") or hmap.get("team name")
    segments = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        market = row[mi] if mi is not None else None
        team = row[ti] if ti is not None else None
        country = rollup_country(str(market or "").strip()) if market else None
        segment = None
        if team:
            parts = str(team).split("-")
            if len(parts) >= 2:
                country = rollup_country(parts[0].strip())
                seg_raw = parts[1].strip().upper()
                if seg_raw in KNOWN:
                    segment = seg_raw
        if country and segment:
            key = segment_market_key(country, segment)
            if key and key not in seen:
                seen.add(key)
                segments.append(
                    {"market": key, "country": country, "segment": segment}
                )
    wb.close()
    return segments


def main() -> None:
    hc = json.loads(HC.read_text())
    json_markets = hc["markets"]
    embedded = expand_embedded(load_embedded())

    rep_segments = rep_rollups_from_xlsx()
    if not rep_segments:
        rep_segments = [
            {"market": k, "country": parse_market_key(k)["country"], "segment": parse_market_key(k)["segment"]}
            for k in load_embedded().keys()
        ]

    primary_by = build_by_key(json_markets)
    embedded_by = build_by_key(embedded)

    filled = 0
    missing = []
    for s in rep_segments:
        w = lookup(primary_by, s["country"], s["segment"], s["market"])
        if not w or w.get("revenue_90d") is None:
            w = lookup(embedded_by, s["country"], s["segment"], s["market"])
        if w and w.get("revenue_90d") is not None:
            filled += 1
        else:
            missing.append(s["market"])

    print(f"Rep segments: {len(rep_segments)}")
    print(f"JSON markets: {len(json_markets)}")
    print(f"Embedded markets: {len(embedded)}")
    print(f"Would fill revenue_90d for {filled}/{len(rep_segments)} segments")
    if missing:
        print("Still missing:", missing)
    usm = lookup(embedded_by, "US", "M", "US-M")
    print(f"US-M embedded revenue: {usm.get('revenue_90d') if usm else None}")


if __name__ == "__main__":
    main()
