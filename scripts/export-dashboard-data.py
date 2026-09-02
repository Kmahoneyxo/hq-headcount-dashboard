#!/usr/bin/env python3
"""Export dashboard JSON sources to CSV and Excel for stakeholders.

Usage:
  python3 scripts/export-dashboard-data.py
  python3 scripts/export-dashboard-data.py path/to/headcount.json

Outputs (by default):
  docs/data/headcount-dashboard.csv           — all market fields (headcount.json)
  docs/data/headcount-dashboard-rep-book.csv  — all reps with book profile (rep_book.json)
  docs/data/headcount-dashboard-book-health.csv — flagged reps (book_health.json)
  docs/data/headcount-dashboard.xlsx          — full workbook (requires openpyxl)
"""

from __future__ import annotations

import csv
import json
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_HC_IN = ROOT / "docs" / "data" / "headcount.json"
DEFAULT_BH_IN = ROOT / "docs" / "data" / "book_health.json"
DEFAULT_RB_IN = ROOT / "docs" / "data" / "rep_book.json"
DEFAULT_IC_IN = ROOT / "docs" / "data" / "impact_coverage_all_reps.json"
OUT_DIR = ROOT / "docs" / "data"
MARKETS_CSV_OUT = OUT_DIR / "headcount-dashboard.csv"
REP_BOOK_CSV_OUT = OUT_DIR / "headcount-dashboard-rep-book.csv"
BOOK_HEALTH_CSV_OUT = OUT_DIR / "headcount-dashboard-book-health.csv"
XLSX_OUT = OUT_DIR / "headcount-dashboard.xlsx"

sys.path.insert(0, str(ROOT / "scripts"))
from build_market_summary import enrich_market  # noqa: E402

# Preferred column order for markets — all dashboard fields, stakeholder-friendly labels.
MARKET_FIELD_LABELS: dict[str, str] = {
    "market": "Market",
    "country": "Country",
    "segment": "Segment",
    "optimal_headcount": "Ideal headcount",
    "optimal_headcount_assigned": "Ideal headcount (assigned accounts)",
    "ideal_pcid": "Ideal PCID (accounts/rep)",
    "avg_pcid_per_rep": "Avg PCID per rep",
    "segment_avg_pcid": "Segment avg PCID",
    "avg_pqr_per_rep": "Avg PQR per rep ($)",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "market_pqr_90d": "Market PQR 90d ($)",
    "rev_vs_pqr_pct": "Rev vs PQR %",
    "current_reps": "Current reps",
    "current_avg_book": "Avg PCID per rep",
    "headcount_gap": "Headcount gap",
    "headcount_recommendation": "HC recommendation",
    "reps_too_big": "Reps too big",
    "reps_too_little": "Reps too little",
    "reps_healthy": "Reps with healthy books",
    "pct_reps_healthy": "% reps healthy",
    "reps_scored": "Reps scored (rep book)",
    "healthy_book_definition": "Healthy book — definition",
    "healthy_book_criteria": "Healthy book — checklist",
    "healthy_book_primary": "Healthy book — summary",
    "splittable_pool": "Splittable PCID pool",
    "total_grow_slots": "Total grow slots",
    "pcid_stddev": "PCID std dev",
    "new_heads_from_split": "New heads from split",
    "book_action": "Book action (Layer 2)",
    "split_hire_recommended": "Split hire recommended",
    "perfect_book_target": "Ideal book size (accounts/rep)",
    "perfect_book_bucket": "Perfect book bucket",
    "perfect_book_ceiling": "Perfect book ceiling",
    "perfect_book_growth_pct": "Perfect book growth %",
    "assigned_accounts": "Assigned accounts",
    "revenue_90d": "Revenue 90d ($)",
    "recommended_action": "Recommended action",
    "avg_pct_book_built": "FY26 % book built",
    "avg_fy26_book_score": "FY26 book score",
    "fy26_target_pct_book_built": "FY26 target % (TBD)",
    "headroom_accounts": "Headroom accounts",
    "sbs_whitespace_country": "SBS whitespace (country)",
    "sbs_whitespace": "SBS whitespace",
    "sbs_revenue_90d": "SBS revenue 90d ($)",
    "books_buildable_from_sbs": "Books buildable from SBS",
    "opp_plateau_book_max": "Opp plateau book max",
    "opp_plateau_rev_per_job": "Opp plateau $/job",
    "opp_pipeline_status": "Opp pipeline status",
    "coverage_inflection_book_max": "Coverage inflection book max",
    "coverage_at_inflection": "Coverage at inflection",
    "median_impact_calls_per_account": "Median impact calls/account",
    "coverage_status": "Coverage status",
    "hc_reason_primary": "HC reason — primary",
    "hc_reason_driver": "HC reason — driver",
    "sbs_has_opportunity": "SBS opportunity",
    "sbs_routing_primary": "SBS routing recommendation",
    "sbs_routing_bullets": "SBS routing — detail",
    "health_primary": "Book health — snapshot",
    "health_bullets": "Book health — detail",
    "recommendation_primary": "Recommendations — top action",
    "recommendation_bullets": "Recommendations — detail",
    "summary_status": "Summary status (HC)",
    "summary_primary": "Book health — snapshot",
    "summary_bullets": "Recommendations — detail",
    "summary_narrative": "Full narrative (health → recs)",
    "optimal_book_primary": "Optimal book — why (plain English)",
    "optimal_book_bullets": "Optimal book — supporting detail",
    "optimal_book_rationale": "Optimal book — full rationale",
    "growth_curve_primary": "Growth curve — narrative",
    "growth_curve_bullets": "Growth curve — detail",
    "growth_peak_accounts": "Growth peak (accounts/rep)",
    "growth_peak_pct": "Growth peak %",
    "growth_decline_above_pcid": "Growth decline above PCIDs",
    "growth_decline_median_pct": "Growth above inflection %",
    "growth_by_bucket": "Growth by PCID bucket (JSON)",
    "segment_avg_jv": "Segment avg JV ($/job)",
    "jv_plateau_book_max": "JV plateau book max",
    "jv_plateau_rev_per_job": "JV plateau $/job",
    "jv_peak_rev_per_job": "JV peak $/job",
    "jv_vs_plateau_pct": "JV vs plateau %",
    "jv_curve_primary": "JV curve — narrative",
    "jv_curve_bullets": "JV curve — detail",
    "jv_by_bucket": "JV by PCID bucket (JSON)",
}

SUMMARY_FIELD_LABELS: dict[str, str] = {
    "market": "Market",
    "country": "Country",
    "segment": "Segment",
    "summary_status": "Status",
    "summary_primary": "HC reason (primary)",
    "hc_reason_primary": "HC reason (primary)",
    "healthy_book_definition": "Healthy book — definition",
    "pct_reps_healthy": "% reps healthy",
    "sbs_routing_primary": "SBS routing",
    "sbs_has_opportunity": "SBS opportunity",
    "summary_bullets": "Supporting detail",
    "optimal_book_primary": "Optimal book — why",
    "optimal_book_rationale": "Optimal book — full rationale",
    "ideal_pcid": "Ideal PCID",
    "perfect_book_target": "Ideal book size",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "headcount_gap": "Headcount gap",
    "headcount_recommendation": "HC recommendation",
    "recommended_action": "Recommended action",
}

SUMMARY_KEY_ORDER = list(SUMMARY_FIELD_LABELS.keys())

FINDINGS_FIELD_LABELS: dict[str, str] = {
    "market": "Market",
    "revenue_90d": "Revenue 90d",
    "current_reps": "Current reps",
    "avg_pcid_per_rep": "Avg PCID",
    "ideal_pcid": "Ideal PCID",
    "optimal_headcount": "Optimal HC",
    "headcount_gap": "HC gap",
    "headcount_recommendation": "HC recommendation",
    "hc_curve_validated": "Curve validated",
    "perfect_book_source": "Perfect book source",
    "growth_peak_accounts": "Growth peak",
    "key_finding": "Key finding (narrative)",
}

FINDINGS_KEY_ORDER = list(FINDINGS_FIELD_LABELS.keys())

MARKET_KEY_ORDER = list(MARKET_FIELD_LABELS.keys())

BOOK_HEALTH_FIELD_LABELS: dict[str, str] = {
    "market": "Market",
    "country": "Country",
    "segment": "Segment",
    "sales_rep_id": "Sales rep ID",
    "sales_team_name": "Sales team",
    "pcid_count": "PCID count",
    "pqr_90d": "PQR 90d ($)",
    "revenue_90d": "Revenue 90d ($)",
    "ideal_pcid": "Ideal PCID",
    "segment_avg_pcid": "Segment avg PCID",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "vs_ideal_pcid": "Vs ideal PCID",
    "too_big": "Too big",
    "too_little": "Too little",
    "peel_to_ideal": "Peel to ideal",
    "grow_slots": "Grow slots",
}

BOOK_HEALTH_KEY_ORDER = list(BOOK_HEALTH_FIELD_LABELS.keys())

REP_BOOK_FIELD_LABELS: dict[str, str] = {
    "market": "Market",
    "country": "Country",
    "segment": "Segment",
    "sales_rep_id": "Sales rep ID",
    "sales_team_name": "Sales team",
    "pcid_count": "PCID count",
    "pqr_90d": "PQR 90d ($)",
    "revenue_90d": "Revenue 90d ($)",
    "impact_calls_90d": "Impact calls 90d",
    "impact_calls_per_account": "Impact calls / account",
    "ideal_pcid": "Ideal PCID",
    "segment_avg_pcid": "Segment avg PCID",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "vs_ideal_pcid": "Vs ideal PCID",
    "too_big": "Too big",
    "too_little": "Too little",
    "healthy_book": "Healthy book",
    "peel_to_ideal": "Peel to ideal",
    "grow_slots": "Grow slots",
}

REP_BOOK_KEY_ORDER = list(REP_BOOK_FIELD_LABELS.keys())

SBS_FIELD_LABELS: dict[str, str] = {
    "country": "Country",
    "segment": "Segment",
    "accounts": "SBS accounts",
    "revenue_90d": "SBS revenue 90d ($)",
}

SBS_KEY_ORDER = list(SBS_FIELD_LABELS.keys())

IMPACT_COVERAGE_FIELD_LABELS: dict[str, str] = {
    "country": "Country",
    "segment": "Segment",
    "sales_rep_id": "Sales rep ID",
    "sales_team_name": "Sales team",
    "pcid_count": "PCID count",
    "impact_calls_90d": "Impact calls 90d",
    "impact_calls_per_account": "Impact calls / account",
    "segment_avg_coverage": "Segment avg coverage",
    "segment_avg_pcid": "Segment avg PCID",
    "segment_avg_pqr": "Segment avg PQR ($)",
    "pqr_90d": "PQR 90d ($)",
    "revenue_90d": "Revenue 90d ($)",
    "too_big_coverage_signal": "Too big (coverage signal)",
}

IMPACT_COVERAGE_KEY_ORDER = list(IMPACT_COVERAGE_FIELD_LABELS.keys())


def market_row(market: dict) -> dict:
    row = dict(market)
    row["market"] = f"{market.get('country', '')}-{market.get('segment', '')}"
    if row.get("sbs_whitespace_country") is None and row.get("sbs_whitespace") is not None:
        row["sbs_whitespace_country"] = row["sbs_whitespace"]
    if not row.get("summary_status"):
        enrich_market(row)
    if isinstance(row.get("growth_by_bucket"), list):
        row["growth_by_bucket"] = json.dumps(row["growth_by_bucket"], separators=(",", ":"))
    if isinstance(row.get("jv_by_bucket"), list):
        row["jv_by_bucket"] = json.dumps(row["jv_by_bucket"], separators=(",", ":"))
    for list_field in (
        "summary_bullets",
        "health_bullets",
        "recommendation_bullets",
        "optimal_book_bullets",
        "sbs_opportunity_bullets",
        "sbs_routing_bullets",
        "impact_coverage_bullets",
        "growth_curve_bullets",
        "jv_curve_bullets",
        "healthy_book_criteria",
    ):
        if isinstance(row.get(list_field), list):
            row[list_field] = " · ".join(row[list_field])
    for key, val in list(row.items()):
        if isinstance(val, list):
            row[key] = " · ".join(str(v) for v in val)
        elif isinstance(val, dict):
            row[key] = " · ".join(f"{k}: {v}" for k, v in val.items())
    return row


def flatten_markets(payload: dict) -> list[dict]:
    return [market_row(m) for m in payload.get("markets", [])]


def findings_row(market: dict) -> dict:
    row = market_row(market)
    row["key_finding"] = (
        row.get("recommendation_primary")
        or row.get("hc_reason_primary")
        or row.get("summary_primary")
        or ""
    )
    if row.get("hc_curve_validated") is False:
        row["hc_curve_validated"] = "No"
    elif row.get("hc_curve_validated") is True:
        row["hc_curve_validated"] = "Yes"
    else:
        row["hc_curve_validated"] = ""
    if row.get("avg_pcid_per_rep") is None and row.get("current_avg_book") is not None:
        row["avg_pcid_per_rep"] = row["current_avg_book"]
    if row.get("ideal_pcid") is None and row.get("perfect_book_target") is not None:
        row["ideal_pcid"] = row["perfect_book_target"]
    return row


def flatten_findings(payload: dict) -> list[dict]:
    rows = [findings_row(m) for m in payload.get("markets", [])]
    rows.sort(key=lambda r: r.get("revenue_90d") or 0, reverse=True)
    return rows


def market_column_keys(markets: list[dict]) -> list[str]:
    seen: set[str] = set()
    keys: list[str] = []
    for key in MARKET_KEY_ORDER:
        if key not in seen:
            keys.append(key)
            seen.add(key)
    for market in markets:
        for key in sorted(market.keys()):
            if key not in seen:
                keys.append(key)
                seen.add(key)
    return keys


def _split_market_key(market_key: str) -> tuple[str, str]:
    if "-" in market_key:
        country, segment = market_key.split("-", 1)
        return country, segment
    return "", ""


def enrich_rep_row(row: dict, market_key: str = "") -> dict:
    out = dict(row)
    if market_key and not out.get("market"):
        out["market"] = market_key
    if not out.get("country") or not out.get("segment"):
        mk = out.get("market") or market_key
        country, segment = _split_market_key(str(mk))
        out.setdefault("country", country)
        out.setdefault("segment", segment)
    return out


def flatten_book_health(payload: dict) -> list[dict]:
    rows: list[dict] = []
    for market_key, market_data in sorted(payload.get("markets", {}).items()):
        for rep in market_data.get("reps", []):
            rows.append(enrich_rep_row(rep, market_key))
    return rows


def flatten_impact_coverage(payload: dict) -> list[dict]:
    reps = payload.get("reps", [])
    rows: list[dict] = []
    for rep in reps:
        row = dict(rep)
        row["market"] = f"{row.get('country', '')}-{row.get('segment', '')}"
        rows.append(row)
    return rows


def flatten_rep_book(payload: dict, book_health: dict | None = None) -> list[dict]:
    reps = payload.get("reps")
    if reps:
        rows = [enrich_rep_row(r) for r in reps]
        for row in rows:
            row["healthy_book"] = not row.get("too_big") and not row.get("too_little")
        return rows
    if book_health:
        return flatten_book_health(book_health)
    return []


def column_keys(rows: list[dict], preferred: list[str]) -> list[str]:
    seen: set[str] = set()
    keys: list[str] = []
    for key in preferred:
        if key not in seen:
            keys.append(key)
            seen.add(key)
    for row in rows:
        for key in sorted(row.keys()):
            if key not in seen:
                keys.append(key)
                seen.add(key)
    return keys


def book_health_column_keys(rows: list[dict]) -> list[str]:
    return column_keys(rows, BOOK_HEALTH_KEY_ORDER)


def rep_book_column_keys(rows: list[dict]) -> list[str]:
    return column_keys(rows, REP_BOOK_KEY_ORDER)


def label_for(key: str, labels: dict[str, str]) -> str:
    return labels.get(key, key.replace("_", " ").title())


def write_csv_rows(
    rows: list[dict],
    keys: list[str],
    labels: dict[str, str],
    path: Path,
    entity_name: str,
) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writerow({k: label_for(k, labels) for k in keys})
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in keys})
    print(f"Wrote {len(rows)} {entity_name} to {path}")


def write_markets_csv(payload: dict, path: Path) -> None:
    markets = flatten_markets(payload)
    keys = market_column_keys(markets)
    write_csv_rows(markets, keys, MARKET_FIELD_LABELS, path, "markets")


def write_book_health_csv(book_health: dict, path: Path) -> None:
    rows = flatten_book_health(book_health)
    keys = book_health_column_keys(rows)
    write_csv_rows(rows, keys, BOOK_HEALTH_FIELD_LABELS, path, "flagged reps")


def write_rep_book_csv(rep_book: dict, book_health: dict, path: Path) -> None:
    rows = flatten_rep_book(rep_book, book_health)
    keys = rep_book_column_keys(rows)
    write_csv_rows(rows, keys, REP_BOOK_FIELD_LABELS, path, "reps")


def style_header_row(ws, bold) -> None:
    for cell in ws[1]:
        cell.font = bold


def append_sheet(wb, title: str, keys: list[str], labels: dict[str, str], rows: list[dict], bold):
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title)
    header = [label_for(k, labels) for k in keys]
    ws.append(header)
    style_header_row(ws, bold)
    for row in rows:
        ws.append([row.get(k) for k in keys])
    for col_idx, key in enumerate(keys, start=1):
        label = label_for(key, labels)
        width = min(max(len(str(label)) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    return ws


def write_findings_sheet(wb, payload: dict, bold) -> None:
    from openpyxl.utils import get_column_letter

    rows = flatten_findings(payload)
    ws = wb.active
    ws.title = "Findings"
    keys = FINDINGS_KEY_ORDER
    ws.append([label_for(k, FINDINGS_FIELD_LABELS) for k in keys])
    style_header_row(ws, bold)
    for row in rows:
        ws.append([row.get(k) for k in keys])
    for col_idx, key in enumerate(keys, start=1):
        label = label_for(key, FINDINGS_FIELD_LABELS)
        width = 18 if key == "key_finding" else min(max(len(str(label)) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.column_dimensions[get_column_letter(len(keys))].width = 72
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def write_xlsx(
    payload: dict,
    book_health: dict,
    rep_book: dict,
    impact_coverage: dict,
    path: Path,
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("openpyxl not installed — skipping .xlsx (pip install openpyxl)")
        return False

    wb = Workbook()
    bold = Font(bold=True)

    write_findings_sheet(wb, payload, bold)

    markets = flatten_markets(payload)
    market_keys = market_column_keys(markets)
    ws_markets = wb.create_sheet("Markets", 1)
    ws_markets.append([label_for(k, MARKET_FIELD_LABELS) for k in market_keys])
    style_header_row(ws_markets, bold)
    for market in markets:
        ws_markets.append([market.get(k) for k in market_keys])
    from openpyxl.utils import get_column_letter

    for col_idx, key in enumerate(market_keys, start=1):
        label = label_for(key, MARKET_FIELD_LABELS)
        width = min(max(len(str(label)) + 2, 12), 36)
        ws_markets.column_dimensions[get_column_letter(col_idx)].width = width
    ws_markets.freeze_panes = "A2"
    ws_markets.auto_filter.ref = ws_markets.dimensions

    rep_rows = flatten_rep_book(rep_book, book_health)
    rep_keys = rep_book_column_keys(rep_rows)
    ws_rep = append_sheet(wb, "Rep book", rep_keys, REP_BOOK_FIELD_LABELS, rep_rows, bold)
    ws_rep.column_dimensions["E"].width = 28

    ic_rows = flatten_impact_coverage(impact_coverage)
    ic_keys = column_keys(ic_rows, IMPACT_COVERAGE_KEY_ORDER)
    ws_ic = append_sheet(wb, "Impact coverage", ic_keys, IMPACT_COVERAGE_FIELD_LABELS, ic_rows, bold)
    ws_ic.column_dimensions["D"].width = 28

    book_rows = flatten_book_health(book_health)
    book_keys = book_health_column_keys(book_rows)
    append_sheet(wb, "Book health (flagged reps)", book_keys, BOOK_HEALTH_FIELD_LABELS, book_rows, bold)

    sbs_rows = payload.get("sbs_whitespace", [])
    append_sheet(wb, "SBS whitespace", SBS_KEY_ORDER, SBS_FIELD_LABELS, sbs_rows, bold)

    summary_rows = flatten_markets(payload)
    ws_sum = append_sheet(
        wb, "Market summaries", SUMMARY_KEY_ORDER, SUMMARY_FIELD_LABELS, summary_rows, bold
    )
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 72
    ws_sum.column_dimensions["F"].width = 88

    ws_meta = wb.create_sheet("About")
    ws_meta.append(["Field", "Value"])
    style_header_row(ws_meta, bold)
    exported_at = date.today().isoformat()
    meta_rows = [
        ("Headcount data as of", payload.get("updated_at", "")),
        ("Revenue window", payload.get("window", "")),
        ("Headcount source query", payload.get("query", "")),
        ("Book health data as of", book_health.get("updated_at", "")),
        ("Book health source query", book_health.get("query", "")),
        ("Book health note", book_health.get("note", "")),
        ("Rep book data as of", rep_book.get("updated_at", "")),
        ("Rep book source query", rep_book.get("query", "")),
        ("Rep book note", rep_book.get("note", "")),
        ("Impact coverage data as of", impact_coverage.get("updated_at", "")),
        ("Impact coverage source query", impact_coverage.get("query", "")),
        ("Impact coverage execution ID", impact_coverage.get("execution_id", "")),
        ("Impact coverage source tables", ", ".join(impact_coverage.get("source_tables", []))),
        ("Impact coverage note", impact_coverage.get("note", "")),
        ("Export generated", exported_at),
        ("Markets in export", len(markets)),
        ("Reps in export (Rep book tab)", len(rep_rows)),
        ("Reps in export (Impact coverage tab)", len(ic_rows)),
        ("Flagged reps in export", len(book_rows)),
        ("SBS whitespace rows", len(sbs_rows)),
    ]
    for label, value in meta_rows:
        ws_meta.append([label, value])
    ws_meta.append([])
    ws_meta.append(["AMER focus markets", ", ".join(payload.get("amer_markets", []))])
    ws_meta.append([])
    ws_meta.append(["Segment note", "Segment = GTM sales segment from team name (M, UMM, ACC, L, NAM, DCA). Grain: country × sales_segment. SBS is country-level only."])
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 72

    wb.save(path)
    print(f"Wrote Excel workbook to {path}")
    return True


def load_json(path: Path) -> dict:
    if not path.is_file():
        print(f"Missing input: {path}")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def load_json_optional(path: Path) -> dict:
    if path.is_file():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def main() -> None:
    hc_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_HC_IN
    bh_path = DEFAULT_BH_IN
    rb_path = DEFAULT_RB_IN
    ic_path = DEFAULT_IC_IN
    if len(sys.argv) > 2:
        bh_path = Path(sys.argv[2])
    if len(sys.argv) > 3:
        rb_path = Path(sys.argv[3])
    if len(sys.argv) > 4:
        ic_path = Path(sys.argv[4])

    payload = load_json(hc_path)
    book_health = load_json(bh_path)
    rep_book = load_json_optional(rb_path)
    impact_coverage = load_json_optional(ic_path)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    write_markets_csv(payload, MARKETS_CSV_OUT)
    write_rep_book_csv(rep_book, book_health, REP_BOOK_CSV_OUT)
    write_book_health_csv(book_health, BOOK_HEALTH_CSV_OUT)
    write_xlsx(payload, book_health, rep_book, impact_coverage, XLSX_OUT)


if __name__ == "__main__":
    main()
