#!/usr/bin/env python3
"""Export PCID market attributes (sql/20) to Excel.

Usage:
  python3 scripts/export-pcid-market-attributes.py

Input:
  docs/data/pcid_market_attributes.json  — Quest prod export

Output:
  docs/data/pcid_market_attributes.xlsx
    - PCID attributes: one row per assigned parent company
    - Market summary: PCID count by country × GTM segment
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "docs" / "data" / "pcid_market_attributes.json"
OUT_PATH = ROOT / "docs" / "data" / "pcid_market_attributes.xlsx"

PCID_FIELDS = [
    ("parent_company_id", "PCID"),
    ("parent_company_name", "Parent name"),
    ("market", "Market"),
    ("country", "Country"),
    ("gtm_segment", "GTM segment"),
    ("dsa_segment", "DSA segment"),
    ("pcid_size_segment", "PCID size segment"),
    ("team", "Team"),
    ("sales_rep_id", "Sales rep ID"),
    ("sales_rep_name", "Sales rep name"),
    ("region", "Region"),
    ("sales_industry_vertical", "Industry vertical"),
    ("billing_country", "Billing country"),
    ("hq_country", "HQ country"),
    ("hq_city", "HQ city"),
    ("hq_state", "HQ state"),
    ("industry", "Industry"),
    ("industry_group", "Industry group"),
    ("industry_sector", "Industry sector"),
    ("sales_business_unit_segment", "Business unit segment"),
    ("is_enterprise", "Enterprise"),
    ("is_staffing_agency", "Staffing agency"),
    ("is_ad_agency", "Ad agency"),
    ("ultimate_parent_id", "Ultimate parent ID"),
    ("ultimate_parent_name", "Ultimate parent name"),
    ("employee_count", "Employee count"),
    ("advertiser_count", "Advertiser count"),
    ("login_count", "Login count"),
    ("account_type", "Account type"),
    ("agency_id", "Agency ID"),
    ("agency_name", "Agency name"),
    ("expansion_account_flag", "Expansion flag"),
    ("torso_account_type", "Torso type"),
    ("torso_status", "Torso status"),
    ("sales_last_impact_covered_date", "Last impact covered"),
]

SUMMARY_FIELDS = [
    ("country", "Country"),
    ("gtm_segment", "GTM segment"),
    ("pcid_count", "PCID count"),
    ("rep_count", "Rep count"),
    ("enterprise_count", "Enterprise count"),
    ("staffing_count", "Staffing agency count"),
]


def load_json(path: Path) -> dict:
    if not path.is_file():
        print(f"Missing input: {path}")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def market_summary(rows: list[dict]) -> list[dict]:
    buckets: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        key = (row.get("country") or "Unknown", row.get("gtm_segment") or "Unknown")
        buckets[key].append(row)

    summary: list[dict] = []
    for (country, segment), group in sorted(buckets.items()):
        reps = {r.get("sales_rep_id") for r in group if r.get("sales_rep_id") is not None}
        summary.append({
            "country": country,
            "gtm_segment": segment,
            "pcid_count": len(group),
            "rep_count": len(reps),
            "enterprise_count": sum(1 for r in group if r.get("is_enterprise")),
            "staffing_count": sum(1 for r in group if r.get("is_staffing_agency")),
        })
    return summary


def write_sheet(ws, fields: list[tuple[str, str]], rows: list[dict], wide_col: str | None = None) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    keys = [k for k, _ in fields]
    labels = [label for _, label in fields]
    ws.append(labels)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(k) for k in keys])
    for col_idx, (_, label) in enumerate(fields, start=1):
        width = min(max(len(label) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if wide_col and wide_col in keys:
        idx = keys.index(wide_col) + 1
        ws.column_dimensions[get_column_letter(idx)].width = 28
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def write_xlsx(rows: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed — pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws_pcid = wb.active
    ws_pcid.title = "PCID attributes"
    write_sheet(ws_pcid, PCID_FIELDS, rows, wide_col="team")

    ws_sum = wb.create_sheet("Market summary")
    write_sheet(ws_sum, SUMMARY_FIELDS, market_summary(rows))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {len(rows)} PCIDs to {path}")


def main() -> None:
    payload = load_json(JSON_PATH)
    rows = payload.get("pcids", payload.get("data", []))
    if not rows and isinstance(payload, list):
        rows = payload
    write_xlsx(rows, OUT_PATH)


if __name__ == "__main__":
    main()
