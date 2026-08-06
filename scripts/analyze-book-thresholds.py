#!/usr/bin/env python3
"""Book-size threshold analysis (Option A) — 85% of peak by country × segment.

Buckets reps three ways: PCID bands (sql/16), std-dev from segment median PCID,
and revenue quartiles. Finds peak bucket and first bucket below 85% of peak for
rev growth, JV, and impact coverage; writes Excel + optional headcount.json patch.

Usage:
  python3 scripts/analyze-book-thresholds.py
  python3 scripts/analyze-book-thresholds.py docs/data/book_threshold_reps.json
  python3 scripts/analyze-book-thresholds.py --update-headcount
"""

from __future__ import annotations

import argparse
import json
import statistics
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from build_market_summary import PCID_BUCKETS, _pcid_bucket  # noqa: E402

DEFAULT_REP_IN = ROOT / "docs" / "data" / "book_threshold_reps.json"
IC_PATH = ROOT / "docs" / "data" / "impact_coverage_all_reps.json"
JV_PATH = ROOT / "docs" / "data" / "rep_jv_all_reps.json"
OUT_XLSX = ROOT / "docs" / "data" / "book_size_threshold_analysis.xlsx"
OUT_SUMMARY = ROOT / "docs" / "data" / "book_threshold_summary.json"
HC_PATH = ROOT / "docs" / "data" / "headcount.json"

THRESHOLD_RATIO = 0.85
MIN_REPS_PER_BUCKET = 5
MIN_PQR = 5000


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "rep_json",
        nargs="?",
        type=Path,
        default=DEFAULT_REP_IN,
        help="Rep-level JSON from sql/22 (default: book_threshold_reps.json)",
    )
    p.add_argument(
        "--update-headcount",
        action="store_true",
        help="Patch threshold_analysis into docs/data/headcount.json",
    )
    p.add_argument(
        "--fallback-merge",
        action="store_true",
        help="Build rep rows from impact_coverage + rep_jv if sql/22 JSON missing",
    )
    return p.parse_args()


def load_reps(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload.get("reps", [])


def merge_fallback_reps() -> list[dict]:
    ic = json.loads(IC_PATH.read_text(encoding="utf-8"))
    jv_lookup: dict[tuple[str, str, int], dict] = {}
    if JV_PATH.is_file():
        jv_payload = json.loads(JV_PATH.read_text(encoding="utf-8"))
        for r in jv_payload.get("reps", []):
            jv_lookup[(r["country"], r["segment"], int(r["sales_rep_id"]))] = r

    reps: list[dict] = []
    for r in ic.get("reps", []):
        prior = r.get("pqr_90d")
        current = r.get("revenue_90d")
        if prior is None or prior < MIN_PQR:
            continue
        growth = max(-0.5, min(1.0, (float(current) - float(prior)) / float(prior)))
        key = (r["country"], r["segment"], int(r["sales_rep_id"]))
        jv = jv_lookup.get(key, {})
        jobs = jv.get("jobs_90d")
        rev_per_job = jv.get("rev_per_job")
        if rev_per_job is None and jobs and current:
            rev_per_job = round(float(current) / float(jobs), 2)
        pcid = r.get("pcid_count") or 0
        impact = r.get("impact_calls_90d") or 0
        reps.append(
            {
                "country": r["country"],
                "segment": r["segment"],
                "sales_rep_id": r["sales_rep_id"],
                "team": r.get("sales_team_name"),
                "pcid_count": pcid,
                "revenue_90d": current,
                "pqr_90d": prior,
                "rev_growth_pct": round(growth, 3),
                "jobs_90d": jobs,
                "rev_per_job": rev_per_job,
                "impact_calls_90d": impact,
                "impact_coverage": round(impact / pcid, 2) if pcid else 0.0,
            }
        )
    return reps


def segment_key(country: str, segment: str) -> tuple[str, str]:
    return country, segment


def group_by_segment(reps: list[dict]) -> dict[tuple[str, str], list[dict]]:
    out: dict[tuple[str, str], list[dict]] = {}
    for r in reps:
        out.setdefault(segment_key(r["country"], r["segment"]), []).append(r)
    return out


def median_or_none(values: list[float]) -> float | None:
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return round(statistics.median(clean), 3 if any(isinstance(v, float) and v != int(v) for v in clean) else 2)


def pcid_bucket_label(pcid: int) -> tuple[int, str, int]:
    order, label, midpoint, upper, _ = _pcid_bucket(pcid)
    return order, label, midpoint


def stddev_bucket(pcid: int, median_pcid: float, std_pcid: float) -> tuple[int, str]:
    if std_pcid <= 0:
        if pcid < median_pcid:
            return 1, "Below median"
        if pcid > median_pcid:
            return 3, "Above median"
        return 2, "At median"
    low = median_pcid - std_pcid
    high = median_pcid + std_pcid
    if pcid < low:
        return 1, f"< median−1σ (< {low:.0f})"
    if pcid < median_pcid:
        return 2, f"median−1σ to median ({low:.0f}–{median_pcid:.0f})"
    if pcid <= high:
        return 3, f"median to median+1σ ({median_pcid:.0f}–{high:.0f})"
    return 4, f"> median+1σ (> {high:.0f})"


def rev_quartile(revenue: float, cuts: list[float]) -> tuple[int, str]:
    if revenue <= cuts[0]:
        return 1, "Q1 (lowest rev)"
    if revenue <= cuts[1]:
        return 2, "Q2"
    if revenue <= cuts[2]:
        return 3, "Q3"
    return 4, "Q4 (highest rev)"


def quartile_cuts(values: list[float]) -> list[float]:
    if len(values) < 4:
        return [0.0, 0.0, 0.0]
    s = sorted(values)
    n = len(s)

    def pct(p: float) -> float:
        idx = (n - 1) * p
        lo = int(idx)
        hi = min(lo + 1, n - 1)
        frac = idx - lo
        return s[lo] * (1 - frac) + s[hi] * frac

    return [pct(0.25), pct(0.50), pct(0.75)]


def aggregate_bucket(
    reps: list[dict],
    bucket_fn,
    bucket_type: str,
) -> list[dict]:
    grouped: dict[tuple, dict] = {}
    for r in reps:
        order, label, extra = bucket_fn(r)
        key = (order, label)
        g = grouped.setdefault(
            key,
            {
                "bucket_order": order,
                "bucket_label": label,
                "bucket_type": bucket_type,
                "pcids": [],
                "growths": [],
                "jvs": [],
                "coverages": [],
            },
        )
        if extra is not None:
            g["bucket_midpoint"] = extra
        g["pcids"].append(r["pcid_count"])
        if r.get("rev_growth_pct") is not None:
            g["growths"].append(r["rev_growth_pct"])
        if r.get("rev_per_job") is not None:
            g["jvs"].append(r["rev_per_job"])
        if r.get("impact_coverage") is not None:
            g["coverages"].append(r["impact_coverage"])

    rows: list[dict] = []
    for key in sorted(grouped, key=lambda k: grouped[k]["bucket_order"]):
        g = grouped[key]
        if len(g["pcids"]) < MIN_REPS_PER_BUCKET:
            continue
        row = {
            "bucket_order": g["bucket_order"],
            "bucket_label": g["bucket_label"],
            "bucket_type": bucket_type,
            "rep_count": len(g["pcids"]),
            "median_pcid": round(statistics.median(g["pcids"])),
            "median_rev_growth_pct": median_or_none(g["growths"]),
            "median_jv": median_or_none(g["jvs"]),
            "median_impact_coverage": median_or_none(g["coverages"]),
        }
        if "bucket_midpoint" in g:
            row["bucket_midpoint"] = g["bucket_midpoint"]
        rows.append(row)
    return rows


def find_peak_and_decline(
    buckets: list[dict],
    metric_key: str,
    order_key: str = "bucket_order",
) -> tuple[dict | None, dict | None, float | None]:
    eligible = [b for b in buckets if b.get("rep_count", 0) >= MIN_REPS_PER_BUCKET]
    if not eligible:
        return None, None, None
    peak = max(eligible, key=lambda b: b.get(metric_key) if b.get(metric_key) is not None else -999)
    peak_val = peak.get(metric_key)
    if peak_val is None:
        return None, None, None
    floor = peak_val * THRESHOLD_RATIO
    sorted_b = sorted(eligible, key=lambda b: b[order_key])
    peak_idx = sorted_b.index(peak)
    decline = None
    for b in sorted_b[peak_idx + 1 :]:
        v = b.get(metric_key)
        if v is not None and v < floor:
            decline = b
            break
    return peak, decline, peak_val


def annotate_buckets(buckets: list[dict]) -> list[dict]:
    metrics = [
        ("median_rev_growth_pct", "growth"),
        ("median_jv", "jv"),
        ("median_impact_coverage", "coverage"),
    ]
    peaks: dict[str, dict | None] = {}
    declines: dict[str, dict | None] = {}
    for mk, short in metrics:
        p, d, _ = find_peak_and_decline(buckets, mk)
        peaks[short] = p
        declines[short] = d

    out: list[dict] = []
    for b in buckets:
        row = dict(b)
        for mk, short in metrics:
            p = peaks[short]
            if p and b["bucket_order"] == p["bucket_order"]:
                row[f"peak_{short}"] = True
            d = declines[short]
            if d and b["bucket_order"] == d["bucket_order"]:
                row[f"decline_{short}_85"] = True
        out.append(row)
    return out


def fmt_pct(p: float | None) -> str:
    if p is None:
        return "—"
    return f"{p * 100:.0f}%"


def fmt_jv(v: float | None) -> str:
    if v is None:
        return "—"
    return f"${v:.2f}/job"


def fmt_cov(v: float | None) -> str:
    if v is None:
        return "—"
    return f"{v:.2f} calls/acct"


def band_short(label: str) -> str:
    if ": " in label:
        return label.split(": ", 1)[1]
    return label


def build_segment_narrative(
    country: str,
    segment: str,
    pcid_buckets: list[dict],
) -> dict:
    growth_peak, growth_decline, growth_peak_val = find_peak_and_decline(
        pcid_buckets, "median_rev_growth_pct"
    )
    jv_peak, jv_decline, jv_peak_val = find_peak_and_decline(pcid_buckets, "median_jv")
    cov_peak, cov_decline, cov_peak_val = find_peak_and_decline(
        pcid_buckets, "median_impact_coverage"
    )

    parts: list[str] = []
    if growth_peak and growth_peak_val is not None:
        band = band_short(growth_peak["bucket_label"])
        parts.append(f"Rev growth peaks at {band} PCIDs ({fmt_pct(growth_peak_val)})")
    if jv_peak and jv_peak_val is not None:
        mid = jv_peak.get("bucket_midpoint") or jv_peak.get("median_pcid")
        parts.append(f"JV peaks at ~{mid} ({fmt_jv(jv_peak_val)})")
    if cov_peak and cov_peak_val is not None:
        mid = cov_peak.get("bucket_midpoint") or cov_peak.get("median_pcid")
        parts.append(f"impact coverage peaks at {mid}")

    # Binding threshold = earliest decline among metrics (smallest book size)
    declines = []
    for name, d, peak in [
        ("growth", growth_decline, growth_peak),
        ("jv", jv_decline, jv_peak),
        ("coverage", cov_decline, cov_peak),
    ]:
        if d is not None:
            declines.append((d["bucket_order"], name, d, peak))

    binding_text = ""
    binding_lo: int | float | None = None
    binding_hi: int | float | None = None
    if declines:
        declines.sort(key=lambda x: x[0])
        bind_order, bind_metric, bind_bucket, bind_peak = declines[0]
        binding_lo = (bind_peak or {}).get("bucket_midpoint") or (bind_peak or {}).get("median_pcid")
        binding_hi = (
            bind_bucket.get("bucket_midpoint")
            or bind_bucket.get("median_pcid")
            or bind_bucket.get("bucket_upper")
        )
        lo, hi = binding_lo, binding_hi
        above_bits: list[str] = []
        if jv_decline and jv_decline.get("median_jv") is not None:
            above_bits.append(f"JV falls to {fmt_jv(jv_decline['median_jv'])}")
        if cov_decline:
            cov_txt = fmt_cov(cov_decline.get("median_impact_coverage"))
            if cov_decline.get("median_impact_coverage") is not None and cov_peak_val is not None:
                if cov_decline["median_impact_coverage"] >= cov_peak_val * 0.95:
                    above_bits.append("coverage flat")
                else:
                    above_bits.append("coverage declining")
            elif cov_decline.get("median_impact_coverage") is not None:
                above_bits.append(f"coverage {cov_txt}")
        if growth_decline and growth_decline.get("median_rev_growth_pct") is not None:
            above_bits.append(f"growth {fmt_pct(growth_decline['median_rev_growth_pct'])}")
        binding_range = f"~{lo}–{hi}" if lo and hi and lo != hi else f"~{hi or lo}"
        binding_text = (
            f"Binding threshold {binding_range} accounts — above 85% drop: "
            + ", ".join(above_bits[:3])
            + "."
        )

    narrative = "; ".join(parts) + "." if parts else ""
    if binding_text:
        narrative = (narrative + " " + binding_text).strip()

    return {
        "market": f"{country}-{segment}",
        "country": country,
        "segment": segment,
        "narrative": narrative,
        "growth_peak_bucket": growth_peak["bucket_label"] if growth_peak else None,
        "growth_peak_pct": growth_peak_val,
        "growth_decline_bucket": growth_decline["bucket_label"] if growth_decline else None,
        "jv_peak_bucket": jv_peak["bucket_label"] if jv_peak else None,
        "jv_peak": jv_peak_val,
        "jv_decline_bucket": jv_decline["bucket_label"] if jv_decline else None,
        "jv_decline": jv_decline.get("median_jv") if jv_decline else None,
        "coverage_peak_bucket": cov_peak["bucket_label"] if cov_peak else None,
        "coverage_peak": cov_peak_val,
        "coverage_decline_bucket": cov_decline["bucket_label"] if cov_decline else None,
        "binding_threshold_low": binding_lo,
        "binding_threshold_high": binding_hi,
    }


def analyze_segment(country: str, segment: str, reps: list[dict]) -> dict:
    pcids = [r["pcid_count"] for r in reps if r.get("pcid_count") is not None]
    median_pcid = statistics.median(pcids) if pcids else 0
    std_pcid = statistics.pstdev(pcids) if len(pcids) > 1 else 0.0
    revs = [float(r["revenue_90d"]) for r in reps if r.get("revenue_90d") is not None]
    cuts = quartile_cuts(revs)

    def pcid_fn(r: dict):
        order, label, midpoint = pcid_bucket_label(int(r.get("pcid_count") or 0))
        return order, label, midpoint

    def std_fn(r: dict):
        order, label = stddev_bucket(int(r.get("pcid_count") or 0), median_pcid, std_pcid)
        return order, label, None

    def rev_fn(r: dict):
        order, label = rev_quartile(float(r.get("revenue_90d") or 0), cuts)
        return order, label, None

    pcid_buckets = annotate_buckets(aggregate_bucket(reps, pcid_fn, "pcid"))
    std_buckets = annotate_buckets(aggregate_bucket(reps, std_fn, "stddev"))
    rev_buckets = annotate_buckets(aggregate_bucket(reps, rev_fn, "rev_quartile"))

    narrative = build_segment_narrative(country, segment, pcid_buckets)
    return {
        "country": country,
        "segment": segment,
        "rep_count": len(reps),
        "segment_median_pcid": round(median_pcid),
        "segment_pcid_stddev": round(std_pcid, 1),
        "pcid_buckets": pcid_buckets,
        "stddev_buckets": std_buckets,
        "rev_quartile_buckets": rev_buckets,
        **narrative,
    }


def rep_detail_rows(reps: list[dict]) -> list[dict]:
    by_seg = group_by_segment(reps)
    cache: dict[tuple[str, str], tuple[float, float, list[float]]] = {}
    rows: list[dict] = []
    for key, seg_reps in sorted(by_seg.items()):
        pcids = [r["pcid_count"] for r in seg_reps]
        median_pcid = statistics.median(pcids)
        std_pcid = statistics.pstdev(pcids) if len(pcids) > 1 else 0.0
        revs = [float(r["revenue_90d"]) for r in seg_reps if r.get("revenue_90d") is not None]
        cuts = quartile_cuts(revs)
        cache[key] = (median_pcid, std_pcid, cuts)

    for r in sorted(reps, key=lambda x: (x["country"], x["segment"], x["sales_rep_id"])):
        key = segment_key(r["country"], r["segment"])
        median_pcid, std_pcid, cuts = cache[key]
        _, pcid_label, _ = pcid_bucket_label(int(r.get("pcid_count") or 0))
        _, std_label = stddev_bucket(int(r.get("pcid_count") or 0), median_pcid, std_pcid)
        _, rev_label = rev_quartile(float(r.get("revenue_90d") or 0), cuts)
        rows.append(
            {
                "country": r["country"],
                "segment": r["segment"],
                "sales_rep_id": r["sales_rep_id"],
                "team": r.get("team"),
                "pcid_count": r.get("pcid_count"),
                "revenue_90d": r.get("revenue_90d"),
                "pqr_90d": r.get("pqr_90d"),
                "rev_growth_pct": r.get("rev_growth_pct"),
                "jobs_90d": r.get("jobs_90d"),
                "rev_per_job": r.get("rev_per_job"),
                "impact_calls_90d": r.get("impact_calls_90d"),
                "impact_coverage": r.get("impact_coverage"),
                "pcid_bucket": pcid_label,
                "stddev_bucket": std_label,
                "rev_quartile": rev_label,
            }
        )
    return rows


def flatten_buckets(segments: list[dict], bucket_key: str) -> list[dict]:
    rows: list[dict] = []
    for seg in segments:
        for b in seg.get(bucket_key, []):
            rows.append(
                {
                    "market": f"{seg['country']}-{seg['segment']}",
                    "country": seg["country"],
                    "segment": seg["segment"],
                    **b,
                }
            )
    return rows


def write_xlsx(
    rep_rows: list[dict],
    segments: list[dict],
    path: Path,
    meta: dict,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl required: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    bold = Font(bold=True)

    def write_sheet(title: str, rows: list[dict], keys: list[str] | None = None):
        ws = wb.create_sheet(title)
        if not rows:
            ws.append(["No data"])
            return
        keys = keys or list(rows[0].keys())
        ws.append(keys)
        for cell in ws[1]:
            cell.font = bold
        for row in rows:
            ws.append([row.get(k) for k in keys])
        for col_idx, key in enumerate(keys, start=1):
            width = min(max(len(str(key)) + 2, 12), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = ws.dimensions

    wb.remove(wb.active)

    rep_keys = [
        "country", "segment", "sales_rep_id", "team", "pcid_count",
        "revenue_90d", "pqr_90d", "rev_growth_pct", "jobs_90d", "rev_per_job",
        "impact_calls_90d", "impact_coverage", "pcid_bucket", "stddev_bucket", "rev_quartile",
    ]
    write_sheet("Rep detail", rep_rows, rep_keys)

    bucket_keys = [
        "market", "country", "segment", "bucket_order", "bucket_label",
        "rep_count", "median_pcid", "median_rev_growth_pct", "median_jv",
        "median_impact_coverage", "peak_growth", "decline_growth_85",
        "peak_jv", "decline_jv_85", "peak_coverage", "decline_coverage_85",
    ]
    for sheet, key in [
        ("PCID buckets", "pcid_buckets"),
        ("Std-dev buckets", "stddev_buckets"),
        ("Rev quartiles", "rev_quartile_buckets"),
    ]:
        flat = flatten_buckets(segments, key)
        normalized = []
        for r in flat:
            normalized.append(
                {
                    "market": r["market"],
                    "country": r["country"],
                    "segment": r["segment"],
                    "bucket_order": r["bucket_order"],
                    "bucket_label": r["bucket_label"],
                    "rep_count": r["rep_count"],
                    "median_pcid": r.get("median_pcid"),
                    "median_rev_growth_pct": r.get("median_rev_growth_pct"),
                    "median_jv": r.get("median_jv"),
                    "median_impact_coverage": r.get("median_impact_coverage"),
                    "peak_growth": r.get("peak_growth"),
                    "decline_growth_85": r.get("decline_growth_85"),
                    "peak_jv": r.get("peak_jv"),
                    "decline_jv_85": r.get("decline_jv_85"),
                    "peak_coverage": r.get("peak_coverage"),
                    "decline_coverage_85": r.get("decline_coverage_85"),
                }
            )
        write_sheet(sheet, normalized, bucket_keys)

    summary_rows = [
        {
            "market": s["market"],
            "rep_count": s["rep_count"],
            "segment_median_pcid": s["segment_median_pcid"],
            "growth_peak_bucket": s.get("growth_peak_bucket"),
            "growth_peak_pct": s.get("growth_peak_pct"),
            "growth_decline_bucket": s.get("growth_decline_bucket"),
            "jv_peak_bucket": s.get("jv_peak_bucket"),
            "jv_peak": s.get("jv_peak"),
            "jv_decline_bucket": s.get("jv_decline_bucket"),
            "jv_decline": s.get("jv_decline"),
            "coverage_peak_bucket": s.get("coverage_peak_bucket"),
            "coverage_decline_bucket": s.get("coverage_decline_bucket"),
            "binding_threshold": f"{s.get('binding_threshold_low')}–{s.get('binding_threshold_high')}",
            "narrative": s.get("narrative"),
        }
        for s in segments
        if s.get("narrative")
    ]
    write_sheet(
        "Segment summary",
        summary_rows,
        list(summary_rows[0].keys()) if summary_rows else None,
    )

    about = wb.create_sheet("About")
    about.append(["Book size threshold analysis (Option A)"])
    about["A1"].font = bold
    for line in [
        "",
        f"Generated: {meta.get('updated_at', date.today().isoformat())}",
        f"Source: {meta.get('query', 'sql/22_book_threshold_analysis.sql')}",
        f"Reps: {meta.get('row_count', len(rep_rows)):,}",
        "",
        "Threshold: metric falls below 85% of segment peak (not 90%).",
        "Windows: current 90d 20260427–20260725 vs prior PQR 20260128–20260426.",
        "PCID buckets: sql/16 bands (1-10 … 126-150, 151+).",
        "Std-dev buckets: vs segment median PCID ± 1σ.",
        "Rev quartiles: Q1–Q4 of revenue_90d within country×segment.",
        "Binding threshold: earliest PCID bucket where any metric drops below 85% of peak.",
    ]:
        about.append([line])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote Excel: {path}")


def patch_headcount(segments: list[dict]) -> None:
    if not HC_PATH.is_file():
        print(f"Skip headcount patch — missing {HC_PATH}")
        return
    hc = json.loads(HC_PATH.read_text(encoding="utf-8"))
    by_market = {s["market"]: s for s in segments}
    for m in hc.get("markets", []):
        key = f"{m['country']}-{m['segment']}"
        s = by_market.get(key)
        if not s:
            continue
        m["threshold_analysis"] = {
            "narrative": s.get("narrative"),
            "growth_peak_bucket": s.get("growth_peak_bucket"),
            "growth_peak_pct": s.get("growth_peak_pct"),
            "jv_peak": s.get("jv_peak"),
            "jv_decline": s.get("jv_decline"),
            "binding_threshold_low": s.get("binding_threshold_low"),
            "binding_threshold_high": s.get("binding_threshold_high"),
        }
    HC_PATH.write_text(json.dumps(hc, indent=2) + "\n", encoding="utf-8")
    print(f"Patched threshold_analysis into {HC_PATH}")


def main() -> None:
    args = parse_args()
    if args.rep_json.is_file():
        reps = load_reps(args.rep_json)
        meta = json.loads(args.rep_json.read_text(encoding="utf-8"))
    elif args.fallback_merge or (IC_PATH.is_file() and JV_PATH.is_file()):
        print(f"Using fallback merge from {IC_PATH.name} + {JV_PATH.name}")
        reps = merge_fallback_reps()
        meta = {
            "updated_at": date.today().isoformat(),
            "query": "merged sql/18 + sql/19",
            "row_count": len(reps),
        }
    else:
        print(f"Missing rep JSON: {args.rep_json}")
        sys.exit(1)

    by_seg = group_by_segment(reps)
    segments = [
        analyze_segment(country, segment, seg_reps)
        for (country, segment), seg_reps in sorted(by_seg.items())
    ]
    rep_rows = rep_detail_rows(reps)

    summary_payload = {
        "updated_at": meta.get("updated_at", date.today().isoformat()),
        "query": meta.get("query", "sql/22_book_threshold_analysis.sql"),
        "execution_id": meta.get("execution_id"),
        "threshold_ratio": THRESHOLD_RATIO,
        "row_count": len(reps),
        "segment_count": len(segments),
        "segments": segments,
    }
    OUT_SUMMARY.parent.mkdir(parents=True, exist_ok=True)
    OUT_SUMMARY.write_text(json.dumps(summary_payload, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote summary: {OUT_SUMMARY} ({len(segments)} segments, {len(reps)} reps)")

    write_xlsx(rep_rows, segments, OUT_XLSX, meta)

    if args.update_headcount:
        patch_headcount(segments)

    us_m = next((s for s in segments if s["country"] == "US" and s["segment"] == "M"), None)
    if us_m:
        print("\nUS-M example:")
        print(us_m["narrative"])


if __name__ == "__main__":
    main()
