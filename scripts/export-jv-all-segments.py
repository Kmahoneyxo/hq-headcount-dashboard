#!/usr/bin/env python3
"""Export JV (job value) from JAM for all reps and segments to Excel.

Usage:
  python3 scripts/export-jv-all-segments.py

Input:
  docs/data/rep_jv_all_reps.json  — sql/19 output

Output:
  docs/data/jv_all_segments.xlsx
    - Rep JV: per-rep jobs, revenue, rev_per_job
    - Segment summary: avg JV by country × segment
"""

from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
JV_PATH = ROOT / "docs" / "data" / "rep_jv_all_reps.json"
OUT_PATH = ROOT / "docs" / "data" / "jv_all_segments.xlsx"

REP_FIELDS = [
    ("sales_rep_id", "Sales rep ID"),
    ("country", "Country"),
    ("segment", "Segment"),
    ("sales_team_name", "Team"),
    ("jobs_90d", "Jobs 90d"),
    ("revenue_90d", "Revenue 90d ($)"),
    ("rev_per_job", "Rev / job (JV)"),
    ("pqr_90d", "Prior Q rev 90d ($)"),
]

SUMMARY_FIELDS = [
    ("country", "Country"),
    ("segment", "Segment"),
    ("rep_count", "Rep count"),
    ("total_jobs_90d", "Total jobs 90d"),
    ("total_revenue_90d", "Total revenue 90d ($)"),
    ("avg_rev_per_job", "Avg rev / job"),
    ("median_rev_per_job", "Median rev / job"),
]


def load_json(path: Path) -> dict:
    if not path.is_file():
        print(f"Missing input: {path}")
        sys.exit(1)
    return json.loads(path.read_text(encoding="utf-8"))


def segment_summary(reps: list[dict]) -> list[dict]:
    buckets: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for rep in reps:
        buckets[(rep["country"], rep["segment"])].append(rep)

    rows: list[dict] = []
    for (country, segment), group in sorted(buckets.items()):
        jobs = sum(int(r.get("jobs_90d") or 0) for r in group)
        revenue = sum(float(r.get("revenue_90d") or 0) for r in group)
        jvs = sorted(float(r["rev_per_job"]) for r in group if r.get("rev_per_job") is not None)
        median = jvs[len(jvs) // 2] if jvs else None
        rows.append({
            "country": country,
            "segment": segment,
            "rep_count": len(group),
            "total_jobs_90d": jobs,
            "total_revenue_90d": round(revenue, 1),
            "avg_rev_per_job": round(revenue / jobs, 2) if jobs else None,
            "median_rev_per_job": median,
        })
    return rows


def write_sheet(ws, fields: list[tuple[str, str]], rows: list[dict], team_col: str | None = None) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    keys = [k for k, _ in fields]
    labels = [label for _, label in fields]
    ws.append(labels)
    for cell in ws[1]:
        cell.font = bold
    for row in rows:
        ws.append([row.get(k) for k in keys])
    for col_idx, (_, label) in enumerate(fields, start=1):
        width = min(max(len(label) + 2, 12), 36)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    if team_col:
        idx = keys.index(team_col) + 1
        ws.column_dimensions[get_column_letter(idx)].width = 28
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions


def write_xlsx(reps: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed — pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    ws_rep = wb.active
    ws_rep.title = "Rep JV"
    write_sheet(ws_rep, REP_FIELDS, reps, team_col="sales_team_name")

    ws_sum = wb.create_sheet("Segment summary")
    summary = segment_summary(reps)
    write_sheet(ws_sum, SUMMARY_FIELDS, summary)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    segments = sorted({r["segment"] for r in reps})
    print(f"Wrote {len(reps)} reps, {len(summary)} country×segment rows to {path}")
    print(f"Segments: {', '.join(segments)}")


def main() -> None:
    payload = load_json(JV_PATH)
    reps = payload.get("reps", [])
    write_xlsx(reps, OUT_PATH)


if __name__ == "__main__":
    main()
