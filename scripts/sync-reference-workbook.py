#!/usr/bin/env python3
"""Compare full reference Excel workbook against dashboard JSON exports.

Save the Google Sheet as Excel (all tabs):
  File → Download → Microsoft Excel (.xlsx)
  → docs/data/reference-workbook.xlsx

Then:
  pip3 install openpyxl   # once
  python3 scripts/sync-reference-workbook.py
  python3 scripts/sync-reference-workbook.py path/to/workbook.xlsx

Writes docs/data/reference_check.json (markets + per-sheet workbook summary).
"""

from __future__ import annotations

import importlib.util
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "docs" / "data" / "reference-workbook.xlsx"
FALLBACK_XLSX = ROOT / "docs" / "data" / "headcount-dashboard.xlsx"
OUT_JSON = ROOT / "docs" / "data" / "reference_check.json"
DATA = ROOT / "docs" / "data"

SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1Hq64TSm77FVH4hLxME2wrs1bJbT8Qi9Puk4FWMdkGsw/edit"
)

MAX_MISMATCH_SAMPLES = 30

# Load label maps from export-dashboard-data.py
_export_spec = importlib.util.spec_from_file_location(
    "export_dashboard_data",
    ROOT / "scripts" / "export-dashboard-data.py",
)
_export = importlib.util.module_from_spec(_export_spec)
_export_spec.loader.exec_module(_export)

MARKET_LABELS = _export.MARKET_FIELD_LABELS
FINDINGS_LABELS = _export.FINDINGS_FIELD_LABELS
SUMMARY_LABELS = _export.SUMMARY_FIELD_LABELS
REP_LABELS = _export.REP_BOOK_FIELD_LABELS
BOOK_LABELS = _export.BOOK_HEALTH_FIELD_LABELS
IC_LABELS = _export.IMPACT_COVERAGE_FIELD_LABELS
SBS_LABELS = _export.SBS_FIELD_LABELS


def labels_to_aliases(labels: dict[str, str]) -> dict[str, list[str]]:
    aliases: dict[str, list[str]] = {}
    for field, label in labels.items():
        aliases[field] = [label.lower(), field.lower().replace("_", " ")]
    return aliases


SHEET_ROLES: list[tuple[str, str, dict[str, list[str]], list[str]]] = [
    ("findings", r"findings|executive", labels_to_aliases(FINDINGS_LABELS), [
        "revenue_90d", "current_reps", "avg_pcid_per_rep", "ideal_pcid",
        "optimal_headcount", "headcount_gap", "headcount_recommendation",
    ]),
    ("markets", r"^markets?$", labels_to_aliases(MARKET_LABELS), [
        "ideal_pcid", "optimal_headcount", "current_reps", "headcount_gap",
        "headcount_recommendation", "assigned_accounts", "avg_pcid_per_rep", "revenue_90d",
    ]),
    ("market_summaries", r"market summar", labels_to_aliases(SUMMARY_LABELS), [
        "ideal_pcid", "headcount_gap", "headcount_recommendation", "pct_reps_healthy",
    ]),
    ("rep_book", r"rep book", labels_to_aliases(REP_LABELS), [
        "pcid_count", "pqr_90d", "revenue_90d", "ideal_pcid", "too_big", "too_little",
        "peel_to_ideal", "grow_slots", "impact_calls_per_account",
    ]),
    ("impact_coverage", r"impact coverage", labels_to_aliases(IC_LABELS), [
        "pcid_count", "impact_calls_90d", "impact_calls_per_account", "pqr_90d", "revenue_90d",
    ]),
    ("book_health", r"book health|flagged", labels_to_aliases(BOOK_LABELS), [
        "pcid_count", "pqr_90d", "ideal_pcid", "too_big", "too_little", "peel_to_ideal", "grow_slots",
    ]),
    ("sbs", r"sbs", labels_to_aliases(SBS_LABELS), ["accounts", "revenue_90d"]),
]

# Global Sales Rep Headcount workbook (Addy / stakeholder model)
GLOBAL_REP_LEVEL_ALIASES = {
    "sales_rep_id": ["rep_id"],
    "pcid_count": ["pcid count"],
    "country": ["market"],
    "sales_team_name": ["team_name"],
    "rep_or_director": ["rep_or_director"],
}

GLOBAL_MODEL_ENGINE_ALIASES = {
    "region": ["region"],
    "avg_pcid": ["average pcid count"],
    "avg_growth": ["average growth"],
    "avg_jv": ["average jv"],
    "recommendation": ["recommendation"],
}

COUNTRY_TO_REGION = {
    "AU": "Asia-Pac", "JP": "Asia-Pac", "IN": "Asia-Pac", "SG": "Asia-Pac",
    "US": "The Americas", "CA": "The Americas", "BR": "The Americas", "MX": "The Americas",
    "UK": "EMEA", "DE": "EMEA", "FR": "EMEA", "NL": "EMEA", "BE": "EMEA", "IE": "EMEA",
    "IT": "EMEA", "ES": "EMEA", "CH": "EMEA", "AT": "EMEA", "PL": "EMEA",
}


def norm_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip().lower())


def map_headers(headers: list[Any], aliases: dict[str, list[str]]) -> dict[str, int]:
    """field -> column index"""
    norm = [norm_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, names in aliases.items():
        for i, h in enumerate(norm):
            if h in names or any(h == n for n in names):
                mapping[field] = i
                break
    return mapping


def parse_bool(raw: Any) -> bool | None:
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip().lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return None


def parse_num(raw: Any) -> float | int | None:
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, bool):
        return int(raw)
    if isinstance(raw, (int, float)):
        f = float(raw)
        return int(f) if f == int(f) else f
    s = str(raw).strip().replace(",", "").replace("$", "").replace("%", "")
    if s.lower() in ("—", "-", "n/a", "na"):
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


def parse_val(field: str, raw: Any) -> Any:
    if field in ("too_big", "too_little", "healthy_book", "too_big_coverage_signal", "sbs_has_opportunity"):
        return parse_bool(raw)
    if field in ("headcount_recommendation", "market", "country", "segment", "sales_team_name", "key_finding"):
        return str(raw).strip() if raw is not None and str(raw).strip() else None
    if field == "hc_curve_validated":
        s = str(raw or "").strip().lower()
        if s in ("yes", "true", "1"):
            return True
        if s in ("no", "false", "0"):
            return False
        return None
    return parse_num(raw)


def norm_rec(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").lower()).strip()


def values_match(field: str, dash: Any, sheet: Any) -> bool:
    if dash is None and sheet is None:
        return True
    if sheet is None:
        return False
    if dash is None:
        return False
    if field == "headcount_recommendation":
        return norm_rec(dash) == norm_rec(sheet)
    if field in ("too_big", "too_little", "healthy_book", "too_big_coverage_signal"):
        return dash == sheet
    d = parse_num(dash)
    s = parse_num(sheet)
    if d is not None and s is not None:
        if isinstance(s, int) and isinstance(d, int):
            return d == s
        return abs(float(d) - float(s)) < 0.5
    return str(dash).strip() == str(sheet).strip()


def row_dict(headers: list[Any], row: tuple, col_map: dict[str, int]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for field, idx in col_map.items():
        if idx < len(row):
            out[field] = parse_val(field, row[idx])
    return out


def market_key_from(row: dict) -> str | None:
    if row.get("market"):
        return str(row["market"]).upper().replace(" ", "")
    c, s = row.get("country"), row.get("segment")
    if c and s:
        return f"{str(c).upper()}-{str(s).upper()}"
    return None


def rep_key_from(row: dict) -> str | None:
    rid = row.get("sales_rep_id")
    if rid is None:
        return None
    m = market_key_from(row)
    return f"{m}:{int(float(rid))}" if m else str(int(float(rid)))


def country_key_from(row: dict) -> str | None:
    c = row.get("country")
    return str(c).upper() if c else None


def is_global_workbook(sheetnames: list[str]) -> bool:
    return any(re.search(r"^rep_level$", n, re.I) for n in sheetnames)


def compare_global_rep_level(ws, dash: dict) -> dict:
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    col_map = map_headers(list(headers), GLOBAL_REP_LEVEL_ALIASES)
    matched = mismatched = missing_dash = sheet_only = 0
    mismatch_samples: list[dict] = []
    sheet_rows = 0
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        sheet_rows += 1
        sheet_row = row_dict(list(headers), row, col_map)
        rid = sheet_row.get("sales_rep_id")
        if rid is None:
            continue
        key = str(int(float(rid)))
        dash_row = dash["rep_by_id"].get(key)
        if not dash_row:
            sheet_only += 1
            if len(mismatch_samples) < MAX_MISMATCH_SAMPLES:
                mismatch_samples.append({"key": key, "issue": "missing_in_dashboard", "sheet_pcid": sheet_row.get("pcid_count")})
            continue
        sv = sheet_row.get("pcid_count")
        dv = dash_row.get("pcid_count")
        if sv is None:
            continue
        if values_match("pcid_count", dv, sv):
            matched += 1
        else:
            mismatched += 1
            if len(mismatch_samples) < MAX_MISMATCH_SAMPLES:
                mismatch_samples.append({
                    "key": key,
                    "diffs": [{"field": "pcid_count", "sheet": sv, "dashboard": dv}],
                })
    return {
        "role": "global_rep_level",
        "name": ws.title,
        "skipped": False,
        "sheet_rows": sheet_rows,
        "dashboard_rows": len(dash["rep_by_id"]),
        "matched": matched,
        "mismatched": mismatched,
        "missing_in_dashboard": sheet_only,
        "columns_mapped": list(col_map.keys()),
        "compare_fields": ["pcid_count"],
        "mismatch_samples": mismatch_samples,
        "note": "Rep_Level PCID Count vs rep_book.json (sql/17). Mismatches often reflect snapshot date or PQR≥$5k filter in dashboard.",
    }


def compare_global_capacity_dashboard(ws, dash: dict) -> dict:
    """Parse pivot-style rep counts per country (col ~9 = market, col ~10 = rep count)."""
    sheet_totals: dict[str, int] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 11:
            continue
        market = row[9]
        count = row[10]
        if market is None or count is None:
            continue
        m = str(market).strip()
        if m in ("Market", "Grand Total", "") or len(m) > 4:
            continue
        n = parse_num(count)
        if n is None:
            continue
        sheet_totals[m.upper()] = int(n)

    matched = mismatched = 0
    mismatch_samples: list[dict] = []
    for country, sheet_reps in sheet_totals.items():
        dash_reps = dash["country_rep_totals"].get(country)
        if dash_reps is None:
            continue
        if int(sheet_reps) == int(dash_reps):
            matched += 1
        else:
            mismatched += 1
            mismatch_samples.append({
                "key": country,
                "diffs": [{"field": "current_reps_sum", "sheet": sheet_reps, "dashboard": dash_reps}],
            })

    return {
        "role": "global_capacity_dashboard",
        "name": ws.title,
        "skipped": False,
        "sheet_rows": len(sheet_totals),
        "dashboard_rows": len(dash["country_rep_totals"]),
        "matched": matched,
        "mismatched": mismatched,
        "missing_in_dashboard": 0,
        "compare_fields": ["current_reps_sum"],
        "mismatch_samples": mismatch_samples,
        "country_totals": sheet_totals,
        "note": "Capacity_Dashboard country rep counts vs sum of current_reps across segments in headcount.json.",
    }


def compare_global_model_engine(ws, dash: dict) -> dict:
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    col_map = map_headers(list(headers), GLOBAL_MODEL_ENGINE_ALIASES)
    matched = mismatched = 0
    mismatch_samples: list[dict] = []
    sheet_rows = 0
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        sheet_rows += 1
        sheet_row = row_dict(list(headers), row, col_map)
        region = sheet_row.get("region")
        if not region:
            continue
        dash_row = dash["region_stats"].get(str(region))
        if not dash_row:
            continue
        ok = True
        diffs = []
        for field in ("avg_pcid", "avg_growth"):
            if field not in col_map:
                continue
            if not values_match(field, dash_row.get(field), sheet_row.get(field)):
                ok = False
                diffs.append({"field": field, "sheet": sheet_row.get(field), "dashboard": dash_row.get(field)})
        if ok:
            matched += 1
        else:
            mismatched += 1
            if len(mismatch_samples) < MAX_MISMATCH_SAMPLES:
                mismatch_samples.append({"key": str(region), "diffs": diffs})
    return {
        "role": "global_model_engine",
        "name": ws.title,
        "skipped": False,
        "sheet_rows": sheet_rows,
        "dashboard_rows": len(dash["region_stats"]),
        "matched": matched,
        "mismatched": mismatched,
        "missing_in_dashboard": 0,
        "compare_fields": ["avg_pcid", "avg_growth"],
        "mismatch_samples": mismatch_samples,
        "note": "Model_Engine regional averages vs rep_book.json rollup (informational — different grain than dashboard markets).",
    }


def detect_role(sheet_name: str, global_mode: bool = False) -> tuple[str, dict[str, list[str]], list[str]] | None:
    low = sheet_name.lower()
    if low == "about":
        return None
    if global_mode:
        if re.match(r"^rep_level$", low):
            return "global_rep_level", GLOBAL_REP_LEVEL_ALIASES, ["pcid_count"]
        if "model_engine" in low:
            return "global_model_engine", GLOBAL_MODEL_ENGINE_ALIASES, ["avg_pcid", "avg_growth"]
        if "capacity_dashboard" in low:
            return "global_capacity", {}, ["current_reps_sum"]
        return None
    for role, pattern, aliases, fields in SHEET_ROLES:
        if re.search(pattern, low):
            return role, aliases, fields
    return None


def load_dashboard() -> dict[str, Any]:
    hc = json.loads((DATA / "headcount.json").read_text())
    bh = json.loads((DATA / "book_health.json").read_text()) if (DATA / "book_health.json").is_file() else {}
    rb = json.loads((DATA / "rep_book.json").read_text()) if (DATA / "rep_book.json").is_file() else {}
    ic = json.loads((DATA / "impact_coverage_all_reps.json").read_text()) if (DATA / "impact_coverage_all_reps.json").is_file() else {}

    markets = [_export.market_row(m) for m in hc.get("markets", [])]
    market_by_key = {market_key_from(m): m for m in markets if market_key_from(m)}

    rep_rows = _export.flatten_rep_book(rb, bh)
    rep_by_key: dict[str, dict] = {}
    rep_by_id: dict[str, dict] = {}
    for r in rep_rows:
        k = rep_key_from(r)
        if k:
            rep_by_key[k] = r
        rid = r.get("sales_rep_id")
        if rid is not None:
            rep_by_id[str(int(rid))] = r

    country_rep_totals: dict[str, int] = {}
    for m in hc.get("markets", []):
        c = str(m.get("country", "")).upper()
        if c:
            country_rep_totals[c] = country_rep_totals.get(c, 0) + int(m.get("current_reps") or 0)

    region_stats: dict[str, dict] = {}
    for r in rep_rows:
        c = str(r.get("country", "")).upper()
        region = COUNTRY_TO_REGION.get(c)
        if not region:
            continue
        bucket = region_stats.setdefault(region, {"pcid_sum": 0, "rep_count": 0, "growth_vals": []})
        pc = r.get("pcid_count")
        if pc is not None:
            bucket["pcid_sum"] += int(pc)
            bucket["rep_count"] += 1
    for region, b in region_stats.items():
        if b["rep_count"]:
            b["avg_pcid"] = b["pcid_sum"] / b["rep_count"]
        else:
            b["avg_pcid"] = None
        b.pop("pcid_sum", None)

    ic_rows = _export.flatten_impact_coverage(ic)
    ic_by_key: dict[str, dict] = {}
    for r in ic_rows:
        k = rep_key_from(r)
        if k:
            ic_by_key[k] = r

    bh_rows = _export.flatten_book_health(bh)
    bh_by_key: dict[str, dict] = {}
    for r in bh_rows:
        k = rep_key_from(r)
        if k:
            bh_by_key[k] = r

    sbs_rows = hc.get("sbs_whitespace", [])
    sbs_by_country = {str(r.get("country", "")).upper(): r for r in sbs_rows if r.get("country")}

    return {
        "headcount": hc,
        "market_by_key": market_by_key,
        "rep_by_key": rep_by_key,
        "rep_by_id": rep_by_id,
        "country_rep_totals": country_rep_totals,
        "region_stats": region_stats,
        "ic_by_key": ic_by_key,
        "bh_by_key": bh_by_key,
        "sbs_by_country": sbs_by_country,
        "summary_rows": _export.flatten_markets(hc),
    }


def dash_lookup(role: str, dash: dict[str, Any], key: str) -> dict | None:
    if role in ("findings", "markets", "market_summaries"):
        return dash["market_by_key"].get(key)
    if role == "rep_book":
        return dash["rep_by_key"].get(key)
    if role == "impact_coverage":
        return dash["ic_by_key"].get(key)
    if role == "book_health":
        return dash["bh_by_key"].get(key)
    if role == "sbs":
        return dash["sbs_by_country"].get(key)
    return None


def row_key(role: str, row: dict) -> str | None:
    if role in ("findings", "markets", "market_summaries"):
        return market_key_from(row)
    if role in ("rep_book", "impact_coverage", "book_health"):
        return rep_key_from(row)
    if role == "sbs":
        return country_key_from(row)
    return None


def compare_sheet(ws, role: str, aliases: dict[str, list[str]], compare_fields: list[str], dash: dict) -> dict:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = list(next(rows_iter))
    except StopIteration:
        return {"role": role, "name": ws.title, "skipped": True, "reason": "empty sheet"}

    col_map = map_headers(list(headers), aliases)
    if not col_map:
        return {"role": role, "name": ws.title, "skipped": True, "reason": "no mapped columns"}

    matched = 0
    mismatched = 0
    missing_dash = 0
    sheet_only = 0
    mismatch_samples: list[dict] = []
    sheet_rows = 0

    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        sheet_rows += 1
        sheet_row = row_dict(list(headers), row, col_map)
        key = row_key(role, sheet_row)
        if not key:
            continue
        dash_row = dash_lookup(role, dash, key)
        if not dash_row:
            missing_dash += 1
            if len(mismatch_samples) < MAX_MISMATCH_SAMPLES:
                mismatch_samples.append({"key": key, "issue": "missing_in_dashboard", "sheet": sheet_row})
            continue

        row_ok = True
        field_diffs = []
        for field in compare_fields:
            if field not in col_map:
                continue
            sv = sheet_row.get(field)
            dv = dash_row.get(field)
            if field == "avg_pcid_per_rep" and dv is None:
                dv = dash_row.get("current_avg_book")
            if field == "ideal_pcid" and dv is None:
                dv = dash_row.get("perfect_book_target")
            if not values_match(field, dv, sv):
                row_ok = False
                field_diffs.append({"field": field, "sheet": sv, "dashboard": dv})

        if row_ok:
            matched += 1
        else:
            mismatched += 1
            if len(mismatch_samples) < MAX_MISMATCH_SAMPLES:
                mismatch_samples.append({"key": key, "diffs": field_diffs})

    dash_count = len(dash["market_by_key"]) if role in ("findings", "markets", "market_summaries") else (
        len(dash["rep_by_key"]) if role == "rep_book" else
        len(dash["ic_by_key"]) if role == "impact_coverage" else
        len(dash["bh_by_key"]) if role == "book_health" else
        len(dash["sbs_by_country"]) if role == "sbs" else 0
    )

    return {
        "role": role,
        "name": ws.title,
        "skipped": False,
        "sheet_rows": sheet_rows,
        "dashboard_rows": dash_count,
        "matched": matched,
        "mismatched": mismatched,
        "missing_in_dashboard": missing_dash,
        "columns_mapped": list(col_map.keys()),
        "compare_fields": [f for f in compare_fields if f in col_map],
        "mismatch_samples": mismatch_samples,
    }


def build_markets_from_workbook(ws_markets, dash: dict) -> dict[str, dict]:
    """Backward-compat markets block from Markets sheet."""
    role_info = detect_role(ws_markets.title)
    if not role_info:
        return {}
    role, aliases, compare_fields = role_info
    rows_iter = ws_markets.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    col_map = map_headers(list(headers), aliases)
    markets: dict[str, dict] = {}
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        sheet_row = row_dict(list(headers), row, col_map)
        key = market_key_from(sheet_row)
        if key:
            markets[key] = sheet_row
    return markets


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.is_file():
        if FALLBACK_XLSX.is_file() and xlsx_path == DEFAULT_XLSX:
            print(f"No {DEFAULT_XLSX} — using local export {FALLBACK_XLSX} for structure test.")
            xlsx_path = FALLBACK_XLSX
        else:
            print(f"Missing workbook: {xlsx_path}")
            print("Download Google Sheet as .xlsx → docs/data/reference-workbook.xlsx")
            print("(Full workbook can be slow — let download finish or copy file from Drive desktop sync.)")
            return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Install openpyxl: pip3 install openpyxl")
        return 1

    dash = load_dashboard()
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    global_mode = is_global_workbook(wb.sheetnames)

    sheet_results: list[dict] = []
    skipped_sheets: list[str] = []

    for name in wb.sheetnames:
        if global_mode:
            low = name.lower()
            if re.match(r"^rep_level$", low):
                result = compare_global_rep_level(wb[name], dash)
                sheet_results.append(result)
            elif "capacity_dashboard" in low:
                result = compare_global_capacity_dashboard(wb[name], dash)
                sheet_results.append(result)
            elif "model_engine" in low:
                result = compare_global_model_engine(wb[name], dash)
                sheet_results.append(result)
            else:
                skipped_sheets.append(name)
                continue
            print(
                f"  {name}: {result.get('matched', 0)} match, {result.get('mismatched', 0)} differ"
                + (f" ({result.get('note', '')[:60]}…)" if result.get("note") else "")
            )
            continue

        detected = detect_role(name, global_mode=False)
        if not detected:
            skipped_sheets.append(name)
            continue
        role, aliases, fields = detected
        result = compare_sheet(wb[name], role, aliases, fields, dash)
        sheet_results.append(result)
        print(
            f"  {name}: {result.get('matched', 0)} match, {result.get('mismatched', 0)} differ"
            + (f", skipped ({result.get('reason')})" if result.get("skipped") else "")
        )

    markets_block: dict[str, dict] = {}
    for name in wb.sheetnames:
        if re.search(r"^markets?$", name.lower()) or re.search(r"findings|executive", name.lower()):
            markets_block = build_markets_from_workbook(wb[name], dash)
            break

    total_matched = sum(s.get("matched", 0) for s in sheet_results if not s.get("skipped"))
    total_mismatched = sum(s.get("mismatched", 0) for s in sheet_results if not s.get("skipped"))

    country_checks: dict[str, dict] = {}
    for s in sheet_results:
        if s.get("role") != "global_capacity_dashboard":
            continue
        for sample in s.get("mismatch_samples", []):
            key = sample.get("key")
            if not key:
                continue
            for d in sample.get("diffs", []):
                if d.get("field") == "current_reps_sum":
                    country_checks[key] = {
                        "sheet_reps": d.get("sheet"),
                        "dashboard_reps": d.get("dashboard"),
                    }
        # also include matched countries from sheet - need to store all in compare function
    for s in sheet_results:
        if s.get("role") == "global_capacity_dashboard" and s.get("country_totals"):
            for country, sheet_reps in s["country_totals"].items():
                dash_reps = dash["country_rep_totals"].get(country)
                country_checks[country] = {
                    "sheet_reps": sheet_reps,
                    "dashboard_reps": dash_reps,
                    "match": int(sheet_reps) == int(dash_reps) if dash_reps is not None else False,
                }

    ref_label = "Global Sales Rep Headcount (1)"
    if xlsx_path.name != "reference-workbook.xlsx":
        ref_label = xlsx_path.stem.strip()

    payload = {
        "source_type": "workbook",
        "workbook_format": "global_headcount" if global_mode else "dashboard_export",
        "reference_label": ref_label,
        "country_checks": country_checks,
        "source_url": SHEET_URL,
        "source_workbook": str(xlsx_path.resolve()),
        "imported_at": date.today().isoformat(),
        "dashboard_snapshot": dash["headcount"].get("updated_at"),
        "workbook_summary": {
            "sheets_in_file": len(wb.sheetnames),
            "sheets_compared": len(sheet_results),
            "sheets_skipped": skipped_sheets,
            "rows_matched": total_matched,
            "rows_mismatched": total_mismatched,
        },
        "sheets": sheet_results,
        "compare_fields": [
            "ideal_pcid", "optimal_headcount", "current_reps", "headcount_gap",
            "headcount_recommendation", "assigned_accounts", "avg_pcid_per_rep",
        ],
        "markets": markets_block,
        "row_count": len(markets_block),
    }

    OUT_JSON.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"Wrote {OUT_JSON} — {total_matched} rows match, {total_mismatched} differ across workbook")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
